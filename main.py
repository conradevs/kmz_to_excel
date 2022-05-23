from zipfile import ZipFile

filename = 'test.kmz'

kmz = ZipFile(filename, 'r')
kml = kmz.open('doc.kml', 'r')
import re
import xml.sax, xml.sax.handler
from kmzHandler import PlacemarkHandler
from Line import Point, Line
from write_excel_cells import excel_fill

parser = xml.sax.make_parser()
handler = PlacemarkHandler()
parser.setContentHandler(handler)
parser.parse(kml)
kmz.close()

def points_lines_shapes(mapping):
    sep = ','
       
    output = 'Name' + sep + 'Coordinates\n'
    #points = ''
    #lines = ''
    #shapes = '' 
    Points = [] 
    Lines = []
    Shapes = []
    for key in mapping:
        coord_str = mapping[key]['coordinates']
        
        coord_str.replace('\n', ',')
        coord_str.replace('\t', ',')
        coord_str.replace('\v', ',')
        coord_str.replace('\f', ',')
        coord_str.replace('\r', ',')
        print(coord_str)
        coord_split = coord_str.split(',')
        #coord_split = re.split('; |, |\*|\n',coord_str)
        #coord_split = re.split('; |, |\*|\n', coord_str)
        #for word in coord_split:
        #    print(word)
        if 'LookAt' in mapping[key]: #points
            #points += key + sep + coord_str + "\n"           
            Points.append(Point(key,coord_split[0],coord_split[1],coord_split[2]))
        elif 'LineString' in mapping[key]: #lines
            #lines += key + sep + coord_str + "\n"
            coord_split = coord_str.split(' ')
            line_points = []
            i=1
            for point in coord_split:
                point_split = point.split(',')
                #print(point)
                if(i>=len(coord_split)): break
                #print(key+'_'+str(i)+' - '+'x: '+point_split[0]+", y: "+point_split[1]+", z: "+point_split[2])
                line_points.append(Point(key+'_'+str(i),point_split[0],point_split[1],point_split[2]))
                i=i+1
            Lines.append(Line(key,line_points))
        else: #shapes
            #shapes += key + sep + coord_str + "\n"
            coord_split = coord_str.split(' ')
            shape_points = []
            i=1
            for point in coord_split:
                point_split = point.split(',')
                #print(point)
                if(i>=len(coord_split)): break
                #print(key+'_'+str(i)+' - '+'x: '+point_split[0]+", y: "+point_split[1]+", z: "+point_split[2])
                shape_points.append(Point(key+'_'+str(i),point_split[0],point_split[1],point_split[2]))
                i=i+1
            Shapes.append(Line(key,shape_points))
    output = [Points,Lines,Shapes]
    return output
 
output = points_lines_shapes(handler.mapping)

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers, is_date_format
from openpyxl.styles.styleable import StyleableObject
from openpyxl.worksheet.hyperlink import Hyperlink

wb = Workbook()
ws =  wb.active
ws.title = "COORDENADAS"
print(output[1][0].name)
for point in output[1][0].points:
    print(point.name+" : " + point.coord_x + ", " + point.coord_y + ", " + point.coord_z)
excel_fill(ws,output[0],output[1],output[2])
out_filename = filename[:-3] + "xlsx" #output filename same as input plus .csv
wb.save(filename = out_filename)


