import sys
from zipfile import ZipFile
import xml.sax, xml.sax.handler
from kmzHandler import PlacemarkHandler
from string_to_objects import points_lines_shapes
from openpyxl import Workbook
from helpers import letter

def excel_fill(ws,points,lines,shapes):
    index=1
    ws['A'+str(index)].value = 'Points'
    index=index+1
    print('index before points: '+str(index))
    for i in range(0,len(points)):
        ws['A'+str(index+i)].value = points[i].name
        ws['B'+str(index+i)].value = points[i].coord_x
        ws['c'+str(index+i)].value = points[i].coord_y
        ws['D'+str(index+i)].value = points[i].coord_z
        ws['F'+str(index+i)].value = "S" + points[i].coord_y[1:10] + "° " + "W"+points[i].coord_x[1:10] + "°"
    index=index+len(points)
    # print('index after points: '+str(index))
    ws['A'+str(index)].value = 'Lines'
    index=index+1
    for i in range(0,len(lines)):
        # print(lines[i].name)
        ws['A'+str(index+i)].value = lines[i].name
        for j in range(1,len(lines[i].points)):
            k=3*(j-1)+1
            if j>=len(lines[i].points): break
            # print("["+letter(k+1)+"],["+letter(k+2)+"],["+letter(k+3)+"]")
            ws[letter(k+1)+str(index+i)].value = lines[i].points[j-1].name
            ws[letter(k+2)+str(index+i)].value = lines[i].points[j-1].coord_x
            ws[letter(k+3)+str(index+i)].value = lines[i].points[j-1].coord_y
            
            #
    index=index+len(lines)
    # print('index after points: '+str(index))
    ws['A'+str(index)].value = 'Shapes'
    index=index+1
    for i in range(0,len(shapes)):
        # print(shapes[i].name)
        ws['A'+str(index+i)].value = shapes[i].name
        for j in range(1,len(shapes[i].points)):
            k=3*(j-1)+1
            if j>=len(shapes[i].points): break
            # print("["+letter(k+1)+"],["+letter(k+2)+"],["+letter(k+3)+"]")
            ws[letter(k+1)+str(index+i)].value = shapes[i].points[j-1].name
            ws[letter(k+2)+str(index+i)].value = shapes[i].points[j-1].coord_x
            ws[letter(k+3)+str(index+i)].value = shapes[i].points[j-1].coord_y


def bdi_ws_fill(ws,point_data):
    ##sorted_points = sorted(point_data, reverse=False,key=lambda point: point.name)
    index=1
    ws['A'+str(index)].value = 'Points'
    index=index+1
    # print('index before points: '+str(index))
    for i in range(0,len(point_data)):
        ws['A'+str(index+i)].value = point_data[i].name
        ws['B'+str(index+i)].value = "S" + point_data[i].coord_y[1:10] + "° " + "W"+point_data[i].coord_x[1:10] + "°"    

def convert_file(file_path_label,output_directory_path_label):

    # print(file_path_label.cget('text'))
    file = file_path_label.cget('text')
    # print(output_directory_path_label.cget('text'))
    output_directory = output_directory_path_label.cget('text')
    kmz = ZipFile(file, 'r')
    kml = kmz.open('doc.kml', 'r')

    parser = xml.sax.make_parser()
    handler = PlacemarkHandler()
    parser.setContentHandler(handler)
    parser.parse(kml)
    kmz.close()

    output = points_lines_shapes(handler.mapping)

    # Create xlsx file
    wb = Workbook()
    #select active tab
    ws =  wb.active
    # name it "COORDENADAS"
    ws.title = "COORDENADAS"
    # call excel cells writing function
    excel_fill(ws,output[0],output[1],output[2])
    # call excel bdi format writing function
    ws_BDI = wb.create_sheet("BDI_LMT")
    bdi_ws_fill(ws_BDI,output[0])
    # output xlsx file same name as kmz file
    out_filename = file.split('/')[-1]
    out_filename = out_filename.split('.')[0]
    out_filename = output_directory + "/" +out_filename + ".xlsx"
    print(out_filename)
    # save output file
    wb.save(filename = out_filename)