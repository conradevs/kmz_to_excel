from msilib.schema import ComboBox
import sys
import copy
from tkinter import *
from zipfile import ZipFile
import xml.sax, xml.sax.handler
from kmzHandler import PlacemarkHandler
from string_to_objects import points_lines_shapes
from openpyxl import load_workbook, Workbook
from Line import Point

BDI_CONFIGURACIONES_INIT = 9
BDI_OPTIONS_INIT = 121

class BDI_Header:
    def __init__(self):
        self.station = ''
        self.circuit = ''
        self.SB_init = ''
        self.SB_end = ''
        self.work_code = ''
        self.conductors_number = ''
        self.line_material = ''
        self.section = ''
        self.cable_isulation = ''
    def load_from_template(self, excel_template):
        wb = load_workbook(filename = excel_template)
        ws = wb.active
        self.station = ws['C4'].value
        self.circuit = ws['M4'].value
        self.voltage = ws['I4'].value
        self.SB_init = ws['C6'].value
        self.SB_end = ws['I6'].value
        self.work_code = ws['I5'].value
        self.conductors_number = ws['C'+str(BDI_CONFIGURACIONES_INIT)].value
        self.line_material = ws['D'+str(BDI_CONFIGURACIONES_INIT)].value
        self.section = ws['E'+str(BDI_CONFIGURACIONES_INIT)].value
        self.cable_isulation = ws['F'+str(BDI_CONFIGURACIONES_INIT)].value
        wb.close
    def fill_form(self,
        work_entry: Entry,
        station_entry: Entry,
        circuit_entry: Entry,
        tension_entry: Entry,
        SB_1_entry: Entry,
        SB_2_entry: Entry,
        conductors_number: ComboBox,
        line_material: ComboBox,
        section: ComboBox,
        cable_isulation: ComboBox,
        output_name: Entry
        ):
        work_entry.delete(0, 'end')
        if(self.work_code != None): work_entry.insert(0,self.work_code)
        station_entry.delete(0, 'end')
        if(self.station != None): station_entry.insert(0,self.station)
        circuit_entry.delete(0, 'end')
        if(self.circuit != None): circuit_entry.insert(0,self.circuit)
        tension_entry.delete(0, 'end')
        if(self.voltage != None): tension_entry.insert(0,self.voltage)
        SB_1_entry.delete(0, 'end')
        if(self.SB_init != None): SB_1_entry.insert(0,self.SB_init)
        SB_2_entry.delete(0, 'end')
        if(self.SB_end != None): SB_2_entry.insert(0,self.SB_end)
        conductors_number.set('')
        if(self.conductors_number != None): conductors_number.set(self.conductors_number)
        line_material.set('')
        if(self.line_material != None): line_material.set(self.line_material)
        section.set('')
        if(self.section != None): section.set(self.section)
        cable_isulation.set('')
        if(self.cable_isulation != None): cable_isulation.set(self.cable_isulation)
        output_name.delete(0, 'end')
        output_name.insert(0,'BDI CPT '+str(self.work_code))

    def get_info_from_form(self,
        work_entry: Entry,
        station_entry: Entry,
        circuit_entry: Entry,
        tension_entry: Entry,
        SB_1_entry: Entry,
        SB_2_entry: Entry,
        conductors_number: ComboBox,
        line_material: ComboBox,
        section: ComboBox,
        cable_isulation: ComboBox
        ):
        self.station = station_entry.get()
        self.circuit = circuit_entry.get()
        self.voltage = tension_entry.get()
        self.SB_init = SB_1_entry.get()
        self.SB_end = SB_2_entry.get()
        self.work_code = work_entry.get()
        self.conductors_number = conductors_number.get()
        self.line_material = line_material.get()
        self.section = section.get()
        self.cable_isulation = cable_isulation.get()
    
    def fill_BDI(self,wb: Workbook):
        ws = wb.active
        ws['C4'].value = self.station
        ws['M4'].value = self.circuit
        ws['I4'].value = self.voltage
        ws['C6'].value = self.SB_init
        ws['I6'].value = self.SB_end
        ws['I5'].value = self.work_code

class Column:
    def __init__(self,name):
        self.name = name
    def load_configuration_from_template(self,index,ws):
        self.coordinates = ws['B'+str(index)].value
        self.conductors = ws['C'+str(index)].value
        self.line_material = ws['D'+str(index)].value
        self.section = ws['E'+str(index)].value
        self.cable_isulation = ws['F'+str(index)].value
        self.column_material = ws['G'+str(index)].value
        self.column_height = ws['H'+str(index)].value
        self.strength = ws['I'+str(index)].value
        self.column_role = ws['J'+str(index)].value
        self.length = ws['K'+str(index)].value
        self.configuration = ws['L'+str(index)].value
        self.cross_material = ws['M'+str(index)].value
        self.insulator_type = ws['N'+str(index)].value
        self.insulator_device = ws['O'+str(index)].value
        self.earth = ws['P'+str(index)].value
        self.state = ws['Q'+str(index)].value
        self.rein = ws['R'+str(index)].value
        self.instalation_date = ws['S'+str(index)].value
        self.bird_device = ws['T'+str(index)].value
        self.other_configuration = ws['U'+str(index)].value
        self.other_cross_material = ws['V'+str(index)].value
        self.other_insulator_type = ws['W'+str(index)].value
        self.other_insulator_device = ws['X'+str(index)].value
        self.contract_number = ws['Y'+str(index)].value
        self.observations = ws['Z'+str(index)].value
        self.elements = ws['AA'+str(index)].value

def BDI_configurations_load():
    wb = load_workbook(filename = 'BDI_configuraciones.xlsx')
    ws = wb.active
    # Load isulation BDI options from template
    line_isulation_options = []
    row = BDI_OPTIONS_INIT
    while (ws['AU'+str(row)].value != None):
        #print(ws['A'+str(row)].value)
        line_isulation_options.append(ws['AU'+str(row)].value)
        row = row + 1
    # Load number of cables BDI options from template
    conductors_number = []
    row = BDI_OPTIONS_INIT
    while (ws['AX'+str(row)].value != None):
        #print(ws['A'+str(row)].value)
        conductors_number.append(ws['AX'+str(row)].value)
        row = row + 1    
    # Load conductor material BDI options from template
    conductor_material = []
    row = BDI_OPTIONS_INIT
    while (ws['AY'+str(row)].value != None):
        #print(ws['A'+str(row)].value)
        conductor_material.append(ws['AY'+str(row)].value)
        row = row + 1   
    # Load conductor section BDI options from template
    conductor_section = []
    row = BDI_OPTIONS_INIT
    while (ws['AZ'+str(row)].value != None):
        #print(ws['A'+str(row)].value)
        conductor_section.append(ws['AZ'+str(row)].value)
        row = row + 1  

    output = []
    output.append(line_isulation_options)
    output.append(conductors_number)
    output.append(conductor_material)
    output.append(conductor_section)

    # Load all template post configurations
    row = BDI_CONFIGURACIONES_INIT
    while (ws['A'+str(row)].value != None):
        #print(ws['A'+str(row)].value)
        output.append(Column(ws['A'+str(row)].value))
        output[-1].load_configuration_from_template(row,ws)
        row = row + 1
    wb.close
    return output

def post_params_fill(name, config, config_index, header: BDI_Header):
    post = copy.deepcopy(config[config_index])
    post.name = name
    post.conductors = header.conductors_number
    post.line_material = header.line_material
    post.section = header.section
    post.cable_isulation = header.cable_isulation
    return post
def column_class_constructor(point: Point,configurations, header: BDI_Header):
    CONFIGURATIONS_START = 4
    post_config = point.name.split(' ')
    post = None
    # Apoyo suspension por defecto
    if(len(post_config) == 1):
        post = post_params_fill(post_config[0],configurations,CONFIGURATIONS_START+0,header) # Cargo parametros desde template (poste C4)
        return post
    if(post_config[1]) == 'SU':
        post = post_params_fill(post_config[0],configurations,CONFIGURATIONS_START+0,header) # Cargo parametros desde template (poste C4)
        if(len(post_config)>2 and (post_config[2] == 'col' or post_config[2] == 'Col')):
            post.column_material = 'Columna de hormigón con orificios'
            post.strength = '3000 N (300 Kg)'
            post.cross_material = 'Metálica'
            post.earth = 'Si'
        if(len(post_config)>2 and post_config[2] == 'C5'): # Esfuerzo C5
            post.strength = 'Clase 5'
    # Apoyo suspension en angulo
    if(post_config[1] == 'SA' and post_config[2] == 'BA'):
        post = post_params_fill(post_config[0],configurations,CONFIGURATIONS_START+1,header)
    if(post_config[1] == 'SA' and post_config[2] == 'RS'):
        post = post_params_fill(post_config[0],configurations,CONFIGURATIONS_START+2,header)
    if(post_config[1] == 'SA' and post_config[2] == 'RD'):
        post = post_params_fill(post_config[0],configurations,CONFIGURATIONS_START+3,header)
    if(post_config[1] == 'SD'):
        post = post_params_fill(post_config[0],configurations,CONFIGURATIONS_START+4,header)
    # Apoyo amarre en angulo    
    if(post_config[1] == 'AA' and post_config[2] == 'BA'):
        post = post_params_fill(post_config[0],configurations,CONFIGURATIONS_START+5,header)
    if(post_config[1] == 'AA' and post_config[2] == 'DE'):
        post = post_params_fill(post_config[0],configurations,CONFIGURATIONS_START+6,header)
    # Apoyo amarre en angulo derivacion
    if(post_config[1] == 'AAD' and post_config[2] == 'BA'):
        post = post_params_fill(post_config[0],configurations,CONFIGURATIONS_START+7,header)
    if(post_config[1] == 'AAD' and post_config[2] == 'DE'):
        post = post_params_fill(post_config[0],configurations,CONFIGURATIONS_START+8,header)
    if(post_config[1] == 'SAD' and post_config[2] == 'BA'):
        post = post_params_fill(post_config[0],configurations,CONFIGURATIONS_START+9,header)
    if(post_config[1] == 'SAD' and post_config[2] == 'RS'):
        post = post_params_fill(post_config[0],configurations,CONFIGURATIONS_START+10,header)
    if(post_config[1] == 'SAD' and post_config[2] == 'RD'):
        post = post_params_fill(post_config[0],configurations,CONFIGURATIONS_START+11,header)
    if(post_config[1] == 'TEA'):
        post = post_params_fill(post_config[0],configurations,CONFIGURATIONS_START+12,header)
    if(post_config[1] == 'TE'):
        post = post_params_fill(post_config[0],configurations,CONFIGURATIONS_START+13,header)
    if('Col' in post_config or 'col' in post_config):
        post = post_params_fill(post_config[0],configurations,CONFIGURATIONS_START+0,header)
        post.column_material = 'Columna de hormigón con orificios'
        post.strength = '3000 N (300 Kg)'
        post.cross_material = 'Metálica'
        post.earth = 'Si'
        if('12/500' in post_config):
            post.column_height = '12'
            post.strength = '5000 N (500 Kg)'
        if('12/800' in post_config):
            post.column_height = '12'
            post.strength = '8000 N (800 Kg)'
            post.rein = 'Sin rienda'
        if('12/1200' in post_config):
            post.column_height = '12'
            post.strength = '12000 N (1200 Kg)'
            post.rein = 'Sin rienda'
        if('12/2000' in post_config):
            post.column_height = '12'
            post.strength = '20000 N (2000 Kg)'
            post.rein = 'Sin rienda'

    return post

def BDI_write_row(point: Column, ws, row):
       ws['A'+str(row)].value = point.name
       ws['B'+str(row)].value = point.coordinates
       ws['C'+str(row)].value = point.conductors
       ws['D'+str(row)].value = point.line_material
       ws['E'+str(row)].value = point.section
       ws['F'+str(row)].value = point.cable_isulation
       ws['G'+str(row)].value = point.column_material
       ws['H'+str(row)].value = point.column_height
       ws['I'+str(row)].value = point.strength
       ws['J'+str(row)].value = point.column_role
       ws['K'+str(row)].value = point.length
       ws['L'+str(row)].value = point.configuration
       ws['M'+str(row)].value = point.cross_material
       ws['N'+str(row)].value = point.insulator_type
       ws['O'+str(row)].value = point.insulator_device
       ws['P'+str(row)].value = point.earth
       ws['Q'+str(row)].value = point.state
       ws['R'+str(row)].value = point.rein
       ws['S'+str(row)].value = point.instalation_date
       ws['T'+str(row)].value = point.bird_device
       ws['U'+str(row)].value = point.other_configuration
       ws['V'+str(row)].value = point.other_cross_material
       ws['W'+str(row)].value = point.other_insulator_type   
       ws['X'+str(row)].value = point.other_insulator_device
       ws['Y'+str(row)].value = point.contract_number
       ws['Z'+str(row)].value = point.observations
       ws['AA'+str(row)].value = point.elements

def test_function(output_name,file_type,header_info: BDI_Header,
    config,
    kmz_file_input,
    directory_output,
    carpeta_get,
    estacion_get,
    salida_get,
    tension_get,
    SB_1_get,
    SB_2_get,
    conductors_number_CBox,
    line_material_CBox,
    section_CBox,
    cable_isulation_CBox
    ):

    print ('Output file name : '+output_name)

    kmz = ZipFile(kmz_file_input, 'r')
    kml = kmz.open('doc.kml', 'r')

    parser = xml.sax.make_parser()
    handler = PlacemarkHandler()
    parser.setContentHandler(handler)
    parser.parse(kml)
    kmz.close()

    points = points_lines_shapes(handler.mapping)[0]

    wb = load_workbook(filename = 'BDI_template.xlsx')
    ws = wb.active
    header_info.get_info_from_form(header_info,carpeta_get,
        estacion_get,
        salida_get,
        tension_get,
        SB_1_get,
        SB_2_get,
        conductors_number_CBox,
        line_material_CBox,
        section_CBox,
        cable_isulation_CBox
    )
    header_info.fill_BDI(header_info,wb)
    postes = []
    for point in points:
        post = column_class_constructor(point,config,header_info)
        if(post != None):
            post.coordinates = "S" + point.coord_y[1:10] + "° " + "W"+point.coord_x[1:10] + "°"
            postes.append(post)
    for row in range(BDI_CONFIGURACIONES_INIT,len(postes)+BDI_CONFIGURACIONES_INIT):
        BDI_write_row(postes[row-BDI_CONFIGURACIONES_INIT], ws, row)
    if (file_type==None or file_type==''): file_type = '.xlsx'
    wb.save(filename = directory_output+'/'+output_name+file_type)